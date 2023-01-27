import os
import logging
import openpyxl
from time import sleep
from openpyxl.styles.borders import Border, Side
import json
from config import Config as cfg
import re
from Lib.EXCEPTION_HANDLER import FileProcessError
from datetime import datetime


BEGIN_ROW_PATTERN = 10
BORDER = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

class Excel:
    """ Базовый класс для работы с Excel """


    def task_kill_excel(self):
        """
        Закрыть вес активные excel файлы
        """
        logging.info('Close excel.exe')
        os.system("2>nul taskkill /f /t /im excel.exe")
        sleep(2)

    def add_data_to_log(self, task, parameters_request,  status):
        """
        Записываем в log.xlsx информацию полученных запросах
        :param task: id запроса
        :param forecast: прогноз
        :param status: статус выполненого задания
        """
        BEGIN_ROW_PATTERN = 10
        BORDER = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        data = {
            1: datetime.today().strftime("%d.%m.%Y %H:%M"),
            2: task,
            3: status,
            4: parameters_request
        }
        logging.info(f'Открываем файл: log.xlsx')
        workbook = openpyxl.load_workbook(cfg.log_xlsx)
        logging.debug('workbook.active')
        worksheet = workbook.active
        end_row = worksheet.max_row
        logging.debug('ws_pattern_total.max_row')
        row = 1
        while True:
            if worksheet.cell(row=row, column=1).value is None:
                break
            row += 1
        for number_column, value in data.items():
            logging.info(f'Вносим значение: {value}')
            worksheet.cell(row=row, column=number_column).value = value
            worksheet.cell(row=row, column=number_column).border = BORDER
        workbook.save(cfg.log_xlsx)
        logging.info('Запись в логфайл добавлена')

    def end_row(self, workbook, column):
        """
        Записываем в лог файл информацию
        :param workbook: объект файла excel
        :param column: индекс столбца
        :return :row - индекс строки
        """
        sheet = workbook.active
        row = 2
        while True:
            if sheet.cell(row=row, column=column).value:
                row += 1
                continue
            return row

    def get_values_column(self, workbook_path):
        """
        Возвращаем все значения из второго столбца
        :param workbook_path: путь до файла файла excel
        :return :data_column - словарь со значениями столбца
        """
        logging.info('Run get_values_column')
        data_column = dict()
        logging.info(f'Открываем файл: {workbook_path}')
        workbook = openpyxl.load_workbook(workbook_path)
        logging.debug('workbook.active')
        worksheet = workbook.active
        logging.debug('ws_pattern_total.max_row')
        max_row = self.end_row(workbook, 3)
        logging.info(f'Кол-во строк: {str(max_row)}')
        for row in range(2, max_row + 1):
            value = str(worksheet.cell(row=row, column=2).value).strip()
            if value not in ['None', '']:
                data_column.update({row: value})
        logging.info('get_values_column - complete!')
        return data_column

    def search_row(self, sheet, date_begin):
        """
        Возвращает индекс строки по значению
        :param sheet: объект листа
        :return : row - индекс строки
        """
        max_row = sheet.max_row
        for row in range(6, max_row + 1):
            if date_begin == sheet.cell(row=row, column=1).value:
                return row

    def excel_to_json(self, workbook_path, json_path, date_begin=None):
        """
        Запись данных из excel в словарик для дальнейшей сериализации
        :param workbook_path: путь до файла excel
        :param date_begin: дата начала периода для поиска данных
        :return : json_file путь до созданного json файла
        """
        try:
            logging.info(f'Начинаю обработку файла {workbook_path}')
            workbook = openpyxl.load_workbook(workbook_path)
            with open(json_path, encoding='UTF-8') as out_file:
                file_data = json.load(out_file)
            for sheet in workbook.worksheets:
                if '04' in sheet.title:
                    logging.info(f'Найден лист {sheet.title}')
                    years = self.search_years(sheet, date_begin)
                    december_value = self.get_december_row(sheet)
                    start_search_month = ''
                    if date_begin:
                        #извлекаем из формата даты 2022-03-02T15:57:25 месяц
                        start_search_month = int(re.findall('-(\d{2})-', date_begin)[0])
                    last_row = december_value - 1
                    start_row = ''
                    start_col = ''
                    #Находим строку с месяцом
                    for column in range(1, sheet.max_column + 1):
                        for row in range(1, sheet.max_row + 1):
                            value = sheet.cell(row=row, column=column).value
                            # записываем индексы строки и столбца с месяцами, что бы с них стартовать перебор
                            if value is None:
                                continue
                            elif 'январь' in value.lower():
                                start_row = row
                                start_col = column
                                break
                        break
                    if start_row == '' or start_col == '':
                        raise FileProcessError('В файле не найдены строки с месяцами')
                    #проходим по словарику с годами
                    for year, year_column in years.items():
                        for row in range(start_row, last_row + 1):
                            month = sheet.cell(row=row, column=start_col).value
                            if date_begin:
                                # проверяем месяц, соотвествует ли он необходимому периоду поиска
                                if start_search_month <= int(cfg.dict_months[month.lower()]):
                                    dict = self.write_in_dict(sheet, row, december_value, month, year, year_column)
                                    file_data['data'].append(dict)
                                else:
                                    continue
                            else:
                                dict = self.write_in_dict(sheet, row, december_value, month, year, year_column)
                                file_data['data'].append(dict)
                    logging.info('Данные из файла извлечены')
                    json_file = os.path.join(cfg.load_dir_IPC, 'file_data.json')
                    json.dump(file_data, open(json_file, mode='w', encoding='utf-8'), indent=4, ensure_ascii=False,
                              default=str)
                    logging.info(f'Файл file_data.json создан')
                    workbook.close()
                    return json_file
            raise FileProcessError('В фале не найден лист c именем "04"')
        except Exception as err:
            raise FileProcessError(f'Ошибка при обработке файла {err}')

    def write_in_dict(self, sheet, row, december_value, month, year, year_column):
        """
        Создание словаря с данными для дальнейшей записи в json
        :param sheet: лист
        :param row: индекс строки
        :param december_value: индекс строки с данными к "декабрю предыдущего года"
        :param month: значение месяца
        :param year: значение года
        :param year_column: индекс столбца с годом
        :return : словарь с данными
        """
        dict = {}
        dict['period'] = month
        dict['time'] = str(year)
        dict['previous_month'] = self.clean_value(sheet.cell(row=row, column=year_column).value)
        dict['previous_year'] = self.clean_value(sheet.cell(row=december_value + 1, column=year_column).value)
        return dict

    def clean_value(self, value):
        """
        Очищает значение ячейки если в нем есть примечание формата 1), например, 103б41)
        """
        if isinstance(value, str):
            if ')' in value and ',' in value:
                clean_data = re.findall(r'\d{1,},\d{,2}', value)[0]
                value = float(clean_data.replace(',','.'))
            elif ',' in value:
                value = float(value.replace(',', '.'))
        return value

    def get_december_row(self, sheet):
        """
        Получить индекс строки со значением "к декабрю предыдущего года"
        :param sheet: лист
        :return : индекс строки
        """
        for column in range(1, sheet.max_column + 1):
            for row in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row, column=column).value
                if value is None:
                    continue
                elif 'к декабрю предыдущего года' in value:
                    logging.info(f'Строка с данными "к декабрю предыдущего года" найдена')
                    return row

    def search_years(self, sheet, date_begin=None):
        """
        Поиск столбцов с периодом на листе "Дефляторы" или листе "год"
        :param sheet: объект Worksheet
        :param date_begin: дата формата 2022-03-02T15:57:25
        :return: словарь вида {год:индекс столбца,}. Например,{2020: 3, 2021: 4, 2022: 5, 2023: 6, 2024: 7}
        """
        # Извлекаем месяц и год из дата формата 2022-03-02T15:57:25
        year = ''
        if date_begin:
            year = int(re.findall('^\d{4}', date_begin)[0])
        for column in range(1, sheet.max_column + 1):
            for row in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row, column=column).value
                if value is None:
                    continue
                elif re.findall(r'^\d\d\d\d$', str(value)):
                    #Если нашли дату, идем по этой строке
                    years = {}
                    for column in range(column, sheet.max_column + 1):
                        value = int(sheet.cell(row=row, column=column).value)
                        #Заканчиваем поиск, когда попадается Nano
                        if date_begin:
                            if year == value or year <= value:
                                years[value] = column
                        else:
                            years[value] = column
                    logging.info(f'Индексы строк с годами собраны')
                    return years
                else:
                    continue

        logging.error(f'На листе {sheet.title} не найден период')
        raise FileProcessError(f'На листе {sheet.title} не найден период')


class OpenWorkbook(Excel):
    """Менеджер контекста"""

    def __init__(self, file_path, read_only=False):
        super().__init__()
        self.file_path = file_path
        self.read_only = read_only
        self.wb = None
        self.ws = None

    def __enter__(self):
        self.get_excel_com()
        logging.debug(f'Открываем файл: "{self.file_path}"')
        self.wb = self.excel_com.Workbooks.Open(self.file_path, None, self.read_only)
        return self.wb

    def __exit__(self, exc_type, exc_val, exc_tb):
        logging.debug(f'Закрываем файл: "{self.file_path}"')
        logging.debug(f'Сохранить изменения: "{not self.read_only}"')
        self.wb.Close(not self.read_only)
